import random
import re
import time

import requests
from bs4 import BeautifulSoup, Comment
# from openpyxl import Workbook
from openpyxl import Workbook
from selenium import webdriver

from 附件下载程序 import public_down
from 预处理 import intaa, _remove_attrs, selectsql, insertsql, gettimestr
from selenium.webdriver.firefox.options import Options

# 打开浏览器
firefox_options = Options()
#firefox_options.add_argument('--headless')
# 创建浏览器对象
wd = webdriver.Firefox(firefox_options=firefox_options)
# wd = webdriver.Firefox()
urllist = ['https://www.cnipa.gov.cn/col/col66/index.html', 'https://www.cnipa.gov.cn/col/col53/index.html']
urlX = "https://www.cnipa.gov.cn/"
datalist = [20200327, 20210731]
mkmlist = ['政策解读', '科新']

timem = time.strftime('%Y%m%d%H%M', time.localtime(time.time()))
for i in range(0, len(urllist)):
    wb = Workbook()
    ws = wb.create_sheet('数据', index=0)
    headData = ['标题', '类别', '发布日期', '全文', 'url']
    for tb in range(1, len(headData) + 1):
        ws.cell(row=1, column=tb, value=headData[tb - 1])
    xy = 2
    url = urllist[i]
    dateed = datalist[i]
    mkm = mkmlist[i]

    for index in range(0, 3):
        if index == 0:
            url0 = url
        else:
            url0 = url + '?uid=669&pageNum=' + str(index + 1)

        wd.get(url0)
        #time.sleep(3)
        # wd.refresh()
        # time.sleep(3)
        # cc = wd.find_element_by_xpath('/html/body/div/div/div/div[4]/div/div[2]/ul')

        soup = BeautifulSoup(wd.page_source, "html.parser")

        ul = soup.find('ul', class_="list clearfix")

        for li in ul.find_all('li'):
            print(li)
            ss = li.find('a').get('href')
            dateop = intaa(li.find('span').text)
            title = li.find('a').text
            title = title.replace("\n", '').replace("\t", '').replace(" ", '')
            fujian = []
            if dateop <= dateed:
                continue
            if title[len(title) - 3:len(title)] == "...":
                title = title[0:len(title) - 3]
                sesql = "SELECT * FROM [dbo].[kx-知识产权] WHERE [标题] LIKE '" + title + "%'"
            else:
                sesql = "SELECT * FROM [dbo].[kx-知识产权] WHERE [标题] = '" + title + "'"
            if selectsql(sesql):
                continue
            jixu = False
            for test in li.find_all('a',
                                    href=re.compile('.*?(pdf|docx|doc|xlsx|xls|rar|zip|jpeg|jpg|png|gif|txt|7z|gz)+$')):
                ysrc = test.get('href')
                href = ysrc
                ysrclis = ysrc.split('/')
                for ccn in ysrclis:
                    ysrc = ccn
                print(href)
                public_down(href, './知产商标专利法律资源应用支持系统/' + mkm + '/' + ysrc)
                test.attrs = {'href': '/datafolder/知产商标专利法律资源应用支持系统/' + mkm + '/' + ysrc}
                fujian.append('/datafolder/知产商标专利法律资源应用支持系统/' + mkm + '/' + ysrc)
                fujian = str(fujian)
                tihuan = re.compile('\'')
                fujian = tihuan.sub('"', fujian)
                title = li.find('a').text
                print(title)
                jixu = True
                ws.cell(xy, 1, title)
                ws.cell(xy, 2, mkm)
                ws.cell(xy, 3, dateop)
                ws.cell(xy, 4,
                        '<div><a href="/datafolder/知产商标专利法律资源应用支持系统/' + mkm + '/' + ysrc + '">' + title + '附件' + '</a></div>')
                ws.cell(xy, 5, href)
#                wb.save('d:/收录收录数据/知产' + mkm + str(timem) + '.xls')
                xy += 1

                in_sql = "INSERT INTO [自收录数据].[dbo].[kx-知识产权] ([标题], [类别], [发布日期], [全文], [url],[附件]) VALUES ('" + title + "','" + mkm + "','" + str(
                    dateop) + "','" + "<div>" + '<div><a href="/datafolder/知产商标专利法律资源应用支持系统/' + mkm + '/' + ysrc + '">' + title + '附件' + '</a></div>' + "</div>" + "','" + href + "','" + fujian + "')"
                insertsql(in_sql)
                # print(in_sql)
            if jixu:
                continue
            title = li.find('a').text
            releaseTime = li.find('span').text
            # print(ss)
            # print(title)
            # if ss == re.compile('http'):
            url1 = ss
            # else:
            #     url1 = url + ss
            wd.get(url1)
            time.sleep(3)
            rese = wd.page_source
            so = BeautifulSoup(rese, "html.parser")
            try:
                dateFb = intaa(so.find('div', class_='article-info clearfix').find('span'))
            except Exception as e:
                dateFb = dateop
            box = so.find('div', class_='article-content cont')
            try:
                box.findAll(text=lambda text: isinstance(text, Comment))
            except Exception as e:
                print('跳过')
                continue
            try:
                title = so.find('h1').text
            except Exception as e:
                title = title
            title = title.replace("\n", '').replace("\t", '').replace(" ", '')
            box = _remove_attrs(box)

            [s.extract() for s in box('meta')]
            for test in box.find_all('a',
                                     href=re.compile(
                                         '.*?(pdf|docx|doc|xlsx|xls|rar|zip|jpeg|jpg|png|gif|txt|7z|gz)+$')):
                ysrc = test.get('href')
                href = ysrc
                ysrclis = ysrc.split('/')
                for ccn in ysrclis:
                    ysrc = ccn
                print(href)
                try:
                    public_down(href, './知产商标专利法律资源应用支持系统/' + mkm + '/' + ysrc)
                except Exception as e:
                    print(e)
                    continue
                test.attrs = {'href': '/datafolder/知产商标专利法律资源应用支持系统/' + mkm + '/' + ysrc}
                fujian.append('/datafolder/知产商标专利法律资源应用支持系统/' + mkm + '/' + ysrc + ";")
            for img in box.find_all('img'):
                ysrc = img.get('src')
                src = urlX + ysrc
                ysrclis = ysrc.split('/')
                for ccn in ysrclis:
                    ysrc = ccn
                try:
                    public_down(src, './知产商标专利法律资源应用支持系统/' + mkm + '/' + ysrc)
                except Exception as e:
                    print(e)
                img.attrs = {'src': '/datafolder/知产商标专利法律资源应用支持系统/' + mkm + '/' + ysrc}
                fujian.append('/datafolder/知产商标专利法律资源应用支持系统/' + mkm + '/' + ysrc)
            strcc = ""
            for bb in box:
                strcc += str(bb)
            ws.cell(xy, 1, title)
            ws.cell(xy, 2, mkm)
            ws.cell(xy, 3, dateop)
            ws.cell(xy, 4, '<div>' + strcc + '</div>')
            ws.cell(xy, 5, url1)
#            wb.save('d:/收录/收录数据/知产' + mkm + str(timem) + '.xls')
            xy += 1
            fujian = str(fujian)
            tihuan = re.compile('\'')
            fujian = tihuan.sub('"', fujian)
            wybz = gettimestr()
            wybz = "zscq" + str(wybz) + str(random.randint(0, 9))
            dateFb = str(dateFb).replace('-', '')
            dateFb = str(dateFb)[0:4] + '.' + str(dateFb)[4:6] + '.' + str(dateFb)[6:8]
            dateop = str(dateop)[0:4] + '.' + str(dateop)[4:6] + '.' + str(dateop)[6:8]
            in_sql = "INSERT INTO [自收录数据].[dbo].[kx-知识产权] ([标题], [类别], [发布日期], [全文], [url],[附件],[唯一标志]) VALUES ('" + title + "','" + mkm + "','" + str(
                dateFb) + "','" + "<div>" + strcc + "</div>" + "','" + url1 + "','" + fujian + "','" + wybz + "')"
            # print(strcc)
            insertsql(in_sql)
            # print(in_sql)
    # if xy > 1:
    #     timem = time.strftime('%Y-%m-%d', time.localtime(time.time()))
    #     message = '万赋杰： ' + '知产商标专利法律资源应用支持系统 ' + mkm + timem + ': 更新了' + str(xy - 2) + '条数据'
    #     requests.get(
    #         'http://lawdoo.com:8038/informapi/api/inform?name=万赋杰&key=asacwda5454wdadadadadw&message=' + message)
    #     requests.get(
    #         'http://lawdoo.com:8038/informapi/api/inform?name=童海&key=asacwda5454wdadadadadw&message=' + message)
    #     requests.get(
    #         'http://lawdoo.com:8038/informapi/api/inform?name=黄晓兰&key=asacwda5454wdadadadadw&message=' + message)
    # if xy > 2:
    #     timem = time.strftime('%Y-%m-%d', time.localtime(time.time()))
    #     message = '万赋杰： ' + '知产商标专利法律资源应用支持系统 ' + mkm + timem + ': 更新了' + str(xy - 2) + '条数据'
    #     requests.get(
    #         'http://lawdoo.com:8038/informapi/api/inform?name=万赋杰&key=asacwda5454wdadadadadw&message=' + message)
wd.quit()
